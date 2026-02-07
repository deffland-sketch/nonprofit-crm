import os
import csv
import io
from datetime import datetime, date, timedelta
from functools import wraps
from decimal import Decimal, InvalidOperation

from flask import Flask, render_template, request, redirect, url_for, flash, Response, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileRequired, FileAllowed
from wtforms import StringField, PasswordField, TextAreaField, DecimalField, DateField, SelectField, BooleanField
from wtforms.validators import DataRequired, Email, Length, Optional, NumberRange, ValidationError
from openpyxl import load_workbook

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///crm.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Session security settings
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

# File upload limits (10MB max)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'

# Rate limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Email configuration
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'noreply@wacharters.org')
mail = Mail(app)

# Token serializer for password reset
def get_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'])

# =============================================================================
# SECURITY CONFIGURATION
# =============================================================================

# Allowed email domains and specific allowed emails
ALLOWED_EMAIL_DOMAINS = ['wacharters.org']
ALLOWED_EMAILS = ['deffland@summitps.org']

def is_email_allowed(email):
    """Check if an email is allowed to register."""
    email = email.lower().strip()
    # Check if email is in the allowed list
    if email in ALLOWED_EMAILS:
        return True
    # Check if email domain is allowed
    domain = email.split('@')[-1] if '@' in email else ''
    return domain in ALLOWED_EMAIL_DOMAINS

def validate_allowed_email(form, field):
    """WTForms validator for allowed emails."""
    if not is_email_allowed(field.data):
        raise ValidationError('Registration is restricted to authorized email addresses only.')


# Role-based access control decorators
def admin_required(f):
    """Decorator to require admin role for a route."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def editor_required(f):
    """Decorator to require editor or admin role for a route."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.can_edit():
            flash('You do not have permission to perform this action.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


# HTTPS enforcement for production
@app.before_request
def enforce_https():
    """Redirect HTTP to HTTPS in production."""
    if os.environ.get('FLASK_ENV') == 'production':
        # Check X-Forwarded-Proto header (set by Railway/load balancer)
        if request.headers.get('X-Forwarded-Proto', 'http') != 'https':
            url = request.url.replace('http://', 'https://', 1)
            return redirect(url, code=301)


# Session timeout enforcement
@app.before_request
def check_session_timeout():
    """Check for session timeout and refresh session on activity."""
    session.permanent = True  # Use PERMANENT_SESSION_LIFETIME

    if current_user.is_authenticated:
        last_activity = session.get('last_activity')
        now = datetime.utcnow()

        if last_activity:
            last_activity_time = datetime.fromisoformat(last_activity)
            if (now - last_activity_time) > timedelta(minutes=30):
                logout_user()
                flash('Your session has expired. Please log in again.', 'info')
                return redirect(url_for('login'))

        session['last_activity'] = now.isoformat()


# Security headers
@app.after_request
def add_security_headers(response):
    """Add security headers to all responses."""
    # Prevent clickjacking
    response.headers['X-Frame-Options'] = 'DENY'
    # Prevent MIME type sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # XSS protection (legacy, but still useful)
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # Referrer policy
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # Content Security Policy
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' https://cdn.jsdelivr.net 'unsafe-inline'; "
        "style-src 'self' https://cdn.jsdelivr.net 'unsafe-inline'; "
        "font-src 'self' https://cdn.jsdelivr.net; "
        "img-src 'self' https://wacharters.org data:; "
        "frame-ancestors 'none';"
    )
    # Permissions policy (modern replacement for Feature-Policy)
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    return response


# Error handler for file too large
@app.errorhandler(413)
def file_too_large(e):
    flash('File is too large. Maximum size is 10MB.', 'danger')
    return redirect(request.referrer or url_for('dashboard'))


# =============================================================================
# DATABASE MODELS
# =============================================================================

class User(UserMixin, db.Model):
    # Role constants
    ROLE_VIEWER = 'viewer'
    ROLE_EDITOR = 'editor'
    ROLE_ADMIN = 'admin'
    ROLES = [ROLE_VIEWER, ROLE_EDITOR, ROLE_ADMIN]

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default=ROLE_VIEWER, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    communications = db.relationship('Communication', backref='logged_by', lazy='dynamic')
    tasks = db.relationship('Task', backref='assigned_to', lazy='dynamic')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def is_admin(self):
        return self.role == self.ROLE_ADMIN

    def is_editor(self):
        return self.role in [self.ROLE_EDITOR, self.ROLE_ADMIN]

    def can_edit(self):
        return self.role in [self.ROLE_EDITOR, self.ROLE_ADMIN]

    def can_delete(self):
        return self.role == self.ROLE_ADMIN

    def can_export(self):
        return self.role == self.ROLE_ADMIN


class Donor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(120))
    phone = db.Column(db.String(20))
    address = db.Column(db.String(200))
    city = db.Column(db.String(100))
    state = db.Column(db.String(50))
    zip_code = db.Column(db.String(20))
    interests = db.Column(db.Text)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    donations = db.relationship('Donation', backref='donor', lazy='dynamic', cascade='all, delete-orphan')
    communications = db.relationship('Communication', backref='donor', lazy='dynamic', cascade='all, delete-orphan')
    tasks = db.relationship('Task', backref='donor', lazy='dynamic', cascade='all, delete-orphan')

    @property
    def full_name(self):
        return f"{self.first_name} {self.last_name}"

    @property
    def total_donated(self):
        return sum(d.amount for d in self.donations) or 0

    @property
    def last_donation(self):
        return self.donations.order_by(Donation.date.desc()).first()


class Donation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    donor_id = db.Column(db.Integer, db.ForeignKey('donor.id'), nullable=False)
    amount = db.Column(db.Numeric(10, 2), nullable=False)
    date = db.Column(db.Date, nullable=False, default=date.today)
    donation_type = db.Column(db.String(50))  # one-time, recurring, in-kind
    campaign = db.Column(db.String(100))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Communication(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    donor_id = db.Column(db.Integer, db.ForeignKey('donor.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    comm_type = db.Column(db.String(50), nullable=False)  # email, phone, meeting, letter
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    subject = db.Column(db.String(200))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    donor_id = db.Column(db.Integer, db.ForeignKey('donor.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    description = db.Column(db.String(500), nullable=False)
    due_date = db.Column(db.Date)
    completed = db.Column(db.Boolean, default=False)
    completed_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class DonorRelationship(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    donor_id = db.Column(db.Integer, db.ForeignKey('donor.id'), nullable=False)
    related_donor_id = db.Column(db.Integer, db.ForeignKey('donor.id'), nullable=False)
    relationship_type = db.Column(db.String(50))  # spouse, family, colleague, friend

    donor = db.relationship('Donor', foreign_keys=[donor_id], backref='relationships_from')
    related_donor = db.relationship('Donor', foreign_keys=[related_donor_id], backref='relationships_to')


class AuditLog(db.Model):
    """Audit log for tracking user actions on sensitive data."""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    action = db.Column(db.String(50), nullable=False)  # login, logout, view, create, edit, delete, export
    resource_type = db.Column(db.String(50))  # donor, donation, communication, etc.
    resource_id = db.Column(db.Integer)
    details = db.Column(db.Text)  # Additional context
    ip_address = db.Column(db.String(45))  # Support IPv6
    user_agent = db.Column(db.String(256))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship('User', backref='audit_logs')

    @classmethod
    def log(cls, action, resource_type=None, resource_id=None, details=None, user_id=None):
        """Create an audit log entry."""
        log_entry = cls(
            user_id=user_id if user_id else (current_user.id if current_user.is_authenticated else None),
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            details=details,
            ip_address=request.headers.get('X-Forwarded-For', request.remote_addr),
            user_agent=request.user_agent.string[:256] if request.user_agent.string else None
        )
        db.session.add(log_entry)
        db.session.commit()
        return log_entry


# =============================================================================
# FORMS
# =============================================================================

class LoginForm(FlaskForm):
    email = StringField('Email', validators=[DataRequired(), Email()])
    password = PasswordField('Password', validators=[DataRequired()])


class RegisterForm(FlaskForm):
    name = StringField('Full Name', validators=[DataRequired(), Length(min=2, max=100)])
    email = StringField('Email', validators=[DataRequired(), Email(), validate_allowed_email])
    password = PasswordField('Password', validators=[DataRequired(), Length(min=8)])


class ForgotPasswordForm(FlaskForm):
    email = StringField('Email', validators=[DataRequired(), Email()])


class ResetPasswordForm(FlaskForm):
    password = PasswordField('New Password', validators=[DataRequired(), Length(min=8)])
    confirm_password = PasswordField('Confirm Password', validators=[DataRequired(), Length(min=8)])


class DonorForm(FlaskForm):
    first_name = StringField('First Name', validators=[DataRequired(), Length(max=50)])
    last_name = StringField('Last Name', validators=[DataRequired(), Length(max=50)])
    email = StringField('Email', validators=[Optional(), Email()])
    phone = StringField('Phone', validators=[Optional(), Length(max=20)])
    address = StringField('Address', validators=[Optional(), Length(max=200)])
    city = StringField('City', validators=[Optional(), Length(max=100)])
    state = StringField('State', validators=[Optional(), Length(max=50)])
    zip_code = StringField('ZIP Code', validators=[Optional(), Length(max=20)])
    interests = TextAreaField('Interests', validators=[Optional()])
    notes = TextAreaField('Notes', validators=[Optional()])


class DonationForm(FlaskForm):
    amount = DecimalField('Amount ($)', validators=[DataRequired(), NumberRange(min=0.01)])
    date = DateField('Date', validators=[DataRequired()])
    donation_type = SelectField('Type', choices=[
        ('one-time', 'One-Time'),
        ('recurring', 'Recurring'),
        ('in-kind', 'In-Kind'),
        ('pledge', 'Pledge')
    ])
    campaign = StringField('Campaign', validators=[Optional(), Length(max=100)])
    notes = TextAreaField('Notes', validators=[Optional()])


class CommunicationForm(FlaskForm):
    comm_type = SelectField('Type', choices=[
        ('email', 'Email'),
        ('phone', 'Phone Call'),
        ('meeting', 'Meeting'),
        ('letter', 'Letter'),
        ('event', 'Event'),
        ('other', 'Other')
    ])
    date = DateField('Date', validators=[DataRequired()])
    subject = StringField('Subject', validators=[Optional(), Length(max=200)])
    notes = TextAreaField('Notes', validators=[Optional()])


class TaskForm(FlaskForm):
    description = StringField('Description', validators=[DataRequired(), Length(max=500)])
    due_date = DateField('Due Date', validators=[Optional()])
    donor_id = SelectField('Related Donor', coerce=int, validators=[Optional()])


class RelationshipForm(FlaskForm):
    related_donor_id = SelectField('Related Donor', coerce=int, validators=[DataRequired()])
    relationship_type = SelectField('Relationship Type', choices=[
        ('spouse', 'Spouse/Partner'),
        ('family', 'Family Member'),
        ('colleague', 'Colleague'),
        ('friend', 'Friend'),
        ('board', 'Board Connection'),
        ('other', 'Other')
    ])


class ImportDonorsForm(FlaskForm):
    file = FileField('Excel File', validators=[
        FileRequired(),
        FileAllowed(['xlsx', 'xls'], 'Excel files only (.xlsx, .xls)')
    ])


class ImportDonationsForm(FlaskForm):
    file = FileField('Excel File', validators=[
        FileRequired(),
        FileAllowed(['xlsx', 'xls'], 'Excel files only (.xlsx, .xls)')
    ])


# =============================================================================
# AUTHENTICATION
# =============================================================================

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute", error_message="Too many login attempts. Please wait a minute before trying again.")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data.lower()).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            AuditLog.log('login', 'user', user.id, f'User {user.email} logged in')
            next_page = request.args.get('next')
            flash('Welcome back!', 'success')
            return redirect(next_page if next_page else url_for('dashboard'))
        AuditLog.log('login_failed', 'user', None, f'Failed login attempt for {form.email.data}')
        flash('Invalid email or password.', 'danger')

    return render_template('login.html', form=form)


@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("3 per minute", error_message="Too many registration attempts. Please wait before trying again.")
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    form = RegisterForm()
    if form.validate_on_submit():
        if User.query.filter_by(email=form.email.data.lower()).first():
            flash('Email already registered.', 'danger')
        else:
            # First user becomes admin automatically
            is_first_user = User.query.count() == 0
            user = User(
                name=form.name.data,
                email=form.email.data.lower(),
                role=User.ROLE_ADMIN if is_first_user else User.ROLE_VIEWER
            )
            user.set_password(form.password.data)
            db.session.add(user)
            db.session.commit()
            if is_first_user:
                flash('Account created as Admin! Please log in.', 'success')
            else:
                flash('Account created! Please log in. An admin can upgrade your role.', 'success')
            return redirect(url_for('login'))

    return render_template('register.html', form=form)


@app.route('/logout')
@login_required
def logout():
    AuditLog.log('logout', 'user', current_user.id, f'User {current_user.email} logged out')
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit("3 per minute")
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    form = ForgotPasswordForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data.lower()).first()
        if user:
            # Generate token
            serializer = get_serializer()
            token = serializer.dumps(user.email, salt='password-reset')

            # Build reset URL
            reset_url = url_for('reset_password', token=token, _external=True)

            # Send email
            try:
                msg = Message(
                    'Password Reset Request - WA Charters CRM',
                    recipients=[user.email]
                )
                msg.body = f'''Hello {user.name},

You requested a password reset for your WA Charters CRM account.

Click the link below to reset your password (valid for 1 hour):
{reset_url}

If you did not request this, please ignore this email.

- WA Charters CRM Team
'''
                mail.send(msg)
                AuditLog.log('password_reset_request', 'user', user.id, f'Password reset requested for {user.email}')
            except Exception as e:
                # Log error but don't reveal to user
                AuditLog.log('password_reset_failed', 'user', user.id, f'Failed to send reset email: {str(e)}')

        # Always show success to prevent email enumeration
        flash('If an account exists with that email, you will receive password reset instructions.', 'info')
        return redirect(url_for('login'))

    return render_template('forgot_password.html', form=form)


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    serializer = get_serializer()
    try:
        email = serializer.loads(token, salt='password-reset', max_age=3600)  # 1 hour expiry
    except SignatureExpired:
        flash('The password reset link has expired. Please request a new one.', 'danger')
        return redirect(url_for('forgot_password'))
    except BadSignature:
        flash('Invalid password reset link.', 'danger')
        return redirect(url_for('forgot_password'))

    user = User.query.filter_by(email=email).first()
    if not user:
        flash('Invalid password reset link.', 'danger')
        return redirect(url_for('forgot_password'))

    form = ResetPasswordForm()
    if form.validate_on_submit():
        if form.password.data != form.confirm_password.data:
            flash('Passwords do not match.', 'danger')
        else:
            user.set_password(form.password.data)
            db.session.commit()
            AuditLog.log('password_reset', 'user', user.id, f'Password reset completed for {user.email}')
            flash('Your password has been reset. Please log in.', 'success')
            return redirect(url_for('login'))

    return render_template('reset_password.html', form=form)


# =============================================================================
# DASHBOARD
# =============================================================================

@app.route('/')
@login_required
def dashboard():
    # Key metrics
    total_donors = Donor.query.count()
    total_donations = db.session.query(db.func.sum(Donation.amount)).scalar() or 0

    # This year's donations
    year_start = date(date.today().year, 1, 1)
    ytd_donations = db.session.query(db.func.sum(Donation.amount)).filter(
        Donation.date >= year_start
    ).scalar() or 0

    # Recent donations
    recent_donations = Donation.query.order_by(Donation.date.desc()).limit(5).all()

    # Upcoming tasks (all users see all tasks)
    upcoming_tasks = Task.query.filter(
        Task.completed == False
    ).order_by(Task.due_date.asc()).limit(5).all()

    # Overdue tasks count
    overdue_count = Task.query.filter(
        Task.completed == False,
        Task.due_date < date.today()
    ).count()

    # Recent communications
    recent_comms = Communication.query.order_by(Communication.date.desc()).limit(5).all()

    # Donors added this month
    month_start = date.today().replace(day=1)
    new_donors_month = Donor.query.filter(Donor.created_at >= month_start).count()

    return render_template('dashboard.html',
        total_donors=total_donors,
        total_donations=total_donations,
        ytd_donations=ytd_donations,
        recent_donations=recent_donations,
        upcoming_tasks=upcoming_tasks,
        overdue_count=overdue_count,
        recent_comms=recent_comms,
        new_donors_month=new_donors_month,
        today=date.today()
    )


# =============================================================================
# DONORS
# =============================================================================

@app.route('/donors')
@login_required
def donors():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)

    query = Donor.query
    if search:
        search_term = f'%{search}%'
        query = query.filter(
            db.or_(
                Donor.first_name.ilike(search_term),
                Donor.last_name.ilike(search_term),
                Donor.email.ilike(search_term)
            )
        )

    donors = query.order_by(Donor.last_name, Donor.first_name).paginate(
        page=page, per_page=20, error_out=False
    )

    return render_template('donors.html', donors=donors, search=search)


@app.route('/donors/new', methods=['GET', 'POST'])
@login_required
@editor_required
def new_donor():
    form = DonorForm()
    if form.validate_on_submit():
        donor = Donor(
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            email=form.email.data,
            phone=form.phone.data,
            address=form.address.data,
            city=form.city.data,
            state=form.state.data,
            zip_code=form.zip_code.data,
            interests=form.interests.data,
            notes=form.notes.data
        )
        db.session.add(donor)
        db.session.commit()
        AuditLog.log('create', 'donor', donor.id, f'Created donor: {donor.full_name}')
        flash(f'Donor {donor.full_name} added successfully!', 'success')
        return redirect(url_for('view_donor', donor_id=donor.id))

    return render_template('donor_form.html', form=form, title='Add New Donor')


@app.route('/donors/<int:donor_id>')
@login_required
def view_donor(donor_id):
    donor = Donor.query.get_or_404(donor_id)
    AuditLog.log('view', 'donor', donor_id, f'Viewed donor: {donor.full_name}')
    donations = donor.donations.order_by(Donation.date.desc()).all()
    communications = donor.communications.order_by(Communication.date.desc()).all()
    tasks = donor.tasks.filter_by(completed=False).order_by(Task.due_date.asc()).all()

    # Get relationships
    relationships = DonorRelationship.query.filter(
        db.or_(
            DonorRelationship.donor_id == donor_id,
            DonorRelationship.related_donor_id == donor_id
        )
    ).all()

    return render_template('donor_view.html',
        donor=donor,
        donations=donations,
        communications=communications,
        tasks=tasks,
        relationships=relationships
    )


@app.route('/donors/<int:donor_id>/edit', methods=['GET', 'POST'])
@login_required
@editor_required
def edit_donor(donor_id):
    donor = Donor.query.get_or_404(donor_id)
    form = DonorForm(obj=donor)

    if form.validate_on_submit():
        form.populate_obj(donor)
        db.session.commit()
        AuditLog.log('edit', 'donor', donor.id, f'Edited donor: {donor.full_name}')
        flash('Donor updated successfully!', 'success')
        return redirect(url_for('view_donor', donor_id=donor.id))

    return render_template('donor_form.html', form=form, title='Edit Donor', donor=donor)


@app.route('/donors/<int:donor_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_donor(donor_id):
    donor = Donor.query.get_or_404(donor_id)
    name = donor.full_name
    AuditLog.log('delete', 'donor', donor_id, f'Deleted donor: {name}')
    db.session.delete(donor)
    db.session.commit()
    flash(f'Donor {name} has been deleted.', 'info')
    return redirect(url_for('donors'))


# =============================================================================
# DONATIONS
# =============================================================================

@app.route('/donors/<int:donor_id>/donations/new', methods=['GET', 'POST'])
@login_required
@editor_required
def new_donation(donor_id):
    donor = Donor.query.get_or_404(donor_id)
    form = DonationForm()

    if request.method == 'GET':
        form.date.data = date.today()

    if form.validate_on_submit():
        donation = Donation(
            donor_id=donor.id,
            amount=form.amount.data,
            date=form.date.data,
            donation_type=form.donation_type.data,
            campaign=form.campaign.data,
            notes=form.notes.data
        )
        db.session.add(donation)
        db.session.commit()
        AuditLog.log('create', 'donation', donation.id, f'Recorded ${donation.amount} donation from {donor.full_name}')
        flash(f'Donation of ${donation.amount} recorded!', 'success')
        return redirect(url_for('view_donor', donor_id=donor.id))

    return render_template('donation_form.html', form=form, donor=donor)


@app.route('/donations')
@login_required
def donations():
    page = request.args.get('page', 1, type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    campaign = request.args.get('campaign')

    query = Donation.query

    if start_date:
        query = query.filter(Donation.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Donation.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if campaign:
        query = query.filter(Donation.campaign.ilike(f'%{campaign}%'))

    donations = query.order_by(Donation.date.desc()).paginate(
        page=page, per_page=20, error_out=False
    )

    # Get unique campaigns for filter dropdown
    campaigns = db.session.query(Donation.campaign).filter(
        Donation.campaign.isnot(None),
        Donation.campaign != ''
    ).distinct().all()
    campaigns = [c[0] for c in campaigns]

    return render_template('donations.html',
        donations=donations,
        campaigns=campaigns,
        filters={'start_date': start_date, 'end_date': end_date, 'campaign': campaign}
    )


@app.route('/donations/<int:donation_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_donation(donation_id):
    donation = Donation.query.get_or_404(donation_id)
    donor_id = donation.donor_id
    AuditLog.log('delete', 'donation', donation_id, f'Deleted ${donation.amount} donation from {donation.donor.full_name}')
    db.session.delete(donation)
    db.session.commit()
    flash('Donation deleted.', 'info')
    return redirect(url_for('view_donor', donor_id=donor_id))


# =============================================================================
# COMMUNICATIONS
# =============================================================================

@app.route('/donors/<int:donor_id>/communications/new', methods=['GET', 'POST'])
@login_required
@editor_required
def new_communication(donor_id):
    donor = Donor.query.get_or_404(donor_id)
    form = CommunicationForm()

    if request.method == 'GET':
        form.date.data = date.today()

    if form.validate_on_submit():
        comm = Communication(
            donor_id=donor.id,
            user_id=current_user.id,
            comm_type=form.comm_type.data,
            date=datetime.combine(form.date.data, datetime.min.time()),
            subject=form.subject.data,
            notes=form.notes.data
        )
        db.session.add(comm)
        db.session.commit()
        flash('Communication logged!', 'success')
        return redirect(url_for('view_donor', donor_id=donor.id))

    return render_template('communication_form.html', form=form, donor=donor)


@app.route('/communications')
@login_required
def communications():
    page = request.args.get('page', 1, type=int)

    comms = Communication.query.order_by(Communication.date.desc()).paginate(
        page=page, per_page=20, error_out=False
    )

    return render_template('communications.html', communications=comms)


# =============================================================================
# TASKS
# =============================================================================

@app.route('/tasks')
@login_required
def tasks():
    show_completed = request.args.get('show_completed', 'false') == 'true'

    # All users see all tasks
    query = Task.query

    if not show_completed:
        query = query.filter_by(completed=False)

    tasks = query.order_by(Task.completed, Task.due_date.asc()).all()

    return render_template('tasks.html', tasks=tasks, show_completed=show_completed, today=date.today())


@app.route('/tasks/new', methods=['GET', 'POST'])
@login_required
@editor_required
def new_task():
    form = TaskForm()

    # Populate donor choices
    donors = Donor.query.order_by(Donor.last_name, Donor.first_name).all()
    form.donor_id.choices = [(0, '-- No specific donor --')] + [(d.id, d.full_name) for d in donors]

    if form.validate_on_submit():
        task = Task(
            user_id=current_user.id,
            description=form.description.data,
            due_date=form.due_date.data,
            donor_id=form.donor_id.data if form.donor_id.data != 0 else None
        )
        db.session.add(task)
        db.session.commit()
        flash('Task created!', 'success')
        return redirect(url_for('tasks'))

    return render_template('task_form.html', form=form)


@app.route('/tasks/<int:task_id>/complete', methods=['POST'])
@login_required
def complete_task(task_id):
    task = Task.query.get_or_404(task_id)
    task.completed = True
    task.completed_at = datetime.utcnow()
    db.session.commit()
    flash('Task marked complete!', 'success')

    # Redirect back to referring page
    return redirect(request.referrer or url_for('tasks'))


@app.route('/tasks/<int:task_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_task(task_id):
    task = Task.query.get_or_404(task_id)
    db.session.delete(task)
    db.session.commit()
    flash('Task deleted.', 'info')
    return redirect(url_for('tasks'))


# =============================================================================
# RELATIONSHIPS
# =============================================================================

@app.route('/donors/<int:donor_id>/relationships/new', methods=['GET', 'POST'])
@login_required
@editor_required
def new_relationship(donor_id):
    donor = Donor.query.get_or_404(donor_id)
    form = RelationshipForm()

    # Populate donor choices (exclude current donor)
    other_donors = Donor.query.filter(Donor.id != donor_id).order_by(Donor.last_name, Donor.first_name).all()
    form.related_donor_id.choices = [(d.id, d.full_name) for d in other_donors]

    if form.validate_on_submit():
        relationship = DonorRelationship(
            donor_id=donor.id,
            related_donor_id=form.related_donor_id.data,
            relationship_type=form.relationship_type.data
        )
        db.session.add(relationship)
        db.session.commit()
        flash('Relationship added!', 'success')
        return redirect(url_for('view_donor', donor_id=donor.id))

    return render_template('relationship_form.html', form=form, donor=donor)


# =============================================================================
# REPORTS
# =============================================================================

@app.route('/reports')
@login_required
def reports():
    # Donation summary by year
    yearly_totals = db.session.query(
        db.func.strftime('%Y', Donation.date).label('year'),
        db.func.sum(Donation.amount).label('total'),
        db.func.count(Donation.id).label('count')
    ).group_by('year').order_by(db.desc('year')).all()

    # Donation summary by campaign
    campaign_totals = db.session.query(
        Donation.campaign,
        db.func.sum(Donation.amount).label('total'),
        db.func.count(Donation.id).label('count')
    ).filter(
        Donation.campaign.isnot(None),
        Donation.campaign != ''
    ).group_by(Donation.campaign).order_by(db.desc('total')).all()

    # Top donors
    top_donors = db.session.query(
        Donor,
        db.func.sum(Donation.amount).label('total')
    ).join(Donation).group_by(Donor.id).order_by(db.desc('total')).limit(10).all()

    # Donation type breakdown
    type_totals = db.session.query(
        Donation.donation_type,
        db.func.sum(Donation.amount).label('total'),
        db.func.count(Donation.id).label('count')
    ).group_by(Donation.donation_type).all()

    return render_template('reports.html',
        yearly_totals=yearly_totals,
        campaign_totals=campaign_totals,
        top_donors=top_donors,
        type_totals=type_totals
    )


# =============================================================================
# EXPORTS
# =============================================================================

@app.route('/export/donors')
@login_required
@admin_required
def export_donors():
    AuditLog.log('export', 'donor', None, f'Exported all donor data ({Donor.query.count()} records)')
    donors = Donor.query.order_by(Donor.last_name, Donor.first_name).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['First Name', 'Last Name', 'Email', 'Phone', 'Address', 'City', 'State', 'ZIP', 'Total Donated', 'Interests', 'Notes'])

    # Data
    for donor in donors:
        writer.writerow([
            donor.first_name,
            donor.last_name,
            donor.email or '',
            donor.phone or '',
            donor.address or '',
            donor.city or '',
            donor.state or '',
            donor.zip_code or '',
            float(donor.total_donated),
            donor.interests or '',
            donor.notes or ''
        ])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment;filename=donors_{date.today()}.csv'}
    )


@app.route('/export/donations')
@login_required
@admin_required
def export_donations():
    AuditLog.log('export', 'donation', None, f'Exported all donation data ({Donation.query.count()} records)')
    donations = Donation.query.order_by(Donation.date.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['Date', 'Donor', 'Amount', 'Type', 'Campaign', 'Notes'])

    # Data
    for d in donations:
        writer.writerow([
            d.date.isoformat(),
            d.donor.full_name,
            float(d.amount),
            d.donation_type or '',
            d.campaign or '',
            d.notes or ''
        ])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment;filename=donations_{date.today()}.csv'}
    )


# =============================================================================
# IMPORTS
# =============================================================================

@app.route('/import')
@login_required
def import_data():
    return render_template('import.html')


@app.route('/import/donors', methods=['GET', 'POST'])
@login_required
@admin_required
def import_donors():
    form = ImportDonorsForm()

    if form.validate_on_submit():
        file = form.file.data
        try:
            wb = load_workbook(filename=io.BytesIO(file.read()))
            ws = wb.active

            # Get headers from first row
            headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]

            # Map common column name variations
            column_map = {
                'first_name': ['first name', 'first', 'firstname', 'first_name'],
                'last_name': ['last name', 'last', 'lastname', 'last_name', 'surname'],
                'email': ['email', 'e-mail', 'email address'],
                'phone': ['phone', 'telephone', 'phone number', 'tel'],
                'address': ['address', 'street', 'street address', 'address1'],
                'city': ['city', 'town'],
                'state': ['state', 'province', 'region'],
                'zip_code': ['zip', 'zip code', 'zipcode', 'postal', 'postal code'],
                'interests': ['interests', 'interest', 'areas of interest'],
                'notes': ['notes', 'note', 'comments', 'comment']
            }

            # Find column indices
            col_indices = {}
            for field, variations in column_map.items():
                for i, header in enumerate(headers):
                    if header in variations:
                        col_indices[field] = i
                        break

            # Check required columns
            if 'first_name' not in col_indices or 'last_name' not in col_indices:
                flash('Excel file must have "First Name" and "Last Name" columns.', 'danger')
                return render_template('import_donors.html', form=form)

            # Import rows
            imported = 0
            skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                first_name = row[col_indices['first_name']] if col_indices.get('first_name') is not None else None
                last_name = row[col_indices['last_name']] if col_indices.get('last_name') is not None else None

                # Skip empty rows
                if not first_name or not last_name:
                    skipped += 1
                    continue

                # Check for existing donor by name and email
                email = row[col_indices['email']] if col_indices.get('email') is not None else None
                existing = None
                if email:
                    existing = Donor.query.filter_by(email=email).first()
                if not existing:
                    existing = Donor.query.filter_by(
                        first_name=str(first_name).strip(),
                        last_name=str(last_name).strip()
                    ).first()

                if existing:
                    skipped += 1
                    continue

                donor = Donor(
                    first_name=str(first_name).strip(),
                    last_name=str(last_name).strip(),
                    email=str(email).strip() if email else None,
                    phone=str(row[col_indices['phone']]).strip() if col_indices.get('phone') is not None and row[col_indices['phone']] else None,
                    address=str(row[col_indices['address']]).strip() if col_indices.get('address') is not None and row[col_indices['address']] else None,
                    city=str(row[col_indices['city']]).strip() if col_indices.get('city') is not None and row[col_indices['city']] else None,
                    state=str(row[col_indices['state']]).strip() if col_indices.get('state') is not None and row[col_indices['state']] else None,
                    zip_code=str(row[col_indices['zip_code']]).strip() if col_indices.get('zip_code') is not None and row[col_indices['zip_code']] else None,
                    interests=str(row[col_indices['interests']]).strip() if col_indices.get('interests') is not None and row[col_indices['interests']] else None,
                    notes=str(row[col_indices['notes']]).strip() if col_indices.get('notes') is not None and row[col_indices['notes']] else None
                )
                db.session.add(donor)
                imported += 1

            db.session.commit()
            flash(f'Successfully imported {imported} donors. Skipped {skipped} rows (empty or duplicates).', 'success')
            return redirect(url_for('donors'))

        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'danger')

    return render_template('import_donors.html', form=form)


@app.route('/import/donations', methods=['GET', 'POST'])
@login_required
@admin_required
def import_donations():
    form = ImportDonationsForm()

    if form.validate_on_submit():
        file = form.file.data
        try:
            wb = load_workbook(filename=io.BytesIO(file.read()))
            ws = wb.active

            # Get headers from first row
            headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]

            # Map common column name variations
            column_map = {
                'donor_name': ['donor', 'donor name', 'name', 'full name', 'fullname'],
                'first_name': ['first name', 'first', 'firstname'],
                'last_name': ['last name', 'last', 'lastname', 'surname'],
                'email': ['email', 'e-mail', 'donor email'],
                'amount': ['amount', 'donation', 'gift', 'donation amount', 'gift amount', '$'],
                'date': ['date', 'donation date', 'gift date', 'received date'],
                'type': ['type', 'donation type', 'gift type', 'payment type'],
                'campaign': ['campaign', 'fund', 'appeal', 'designation'],
                'notes': ['notes', 'note', 'comments', 'memo']
            }

            # Find column indices
            col_indices = {}
            for field, variations in column_map.items():
                for i, header in enumerate(headers):
                    if header in variations:
                        col_indices[field] = i
                        break

            # Check required columns
            if 'amount' not in col_indices:
                flash('Excel file must have an "Amount" column.', 'danger')
                return render_template('import_donations.html', form=form)

            has_donor_name = 'donor_name' in col_indices
            has_split_name = 'first_name' in col_indices and 'last_name' in col_indices
            has_email = 'email' in col_indices

            if not has_donor_name and not has_split_name and not has_email:
                flash('Excel file must have donor identification (Name, First/Last Name, or Email).', 'danger')
                return render_template('import_donations.html', form=form)

            # Import rows
            imported = 0
            skipped = 0
            errors = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Parse amount
                    amount_val = row[col_indices['amount']] if col_indices.get('amount') is not None else None
                    if not amount_val:
                        skipped += 1
                        continue

                    # Handle amount formatting (remove $, commas, etc.)
                    if isinstance(amount_val, str):
                        amount_val = amount_val.replace('$', '').replace(',', '').strip()
                    try:
                        amount = Decimal(str(amount_val))
                    except InvalidOperation:
                        errors.append(f'Row {row_num}: Invalid amount "{amount_val}"')
                        continue

                    # Find donor
                    donor = None

                    # Try email first
                    if has_email and row[col_indices['email']]:
                        email = str(row[col_indices['email']]).strip()
                        donor = Donor.query.filter_by(email=email).first()

                    # Try full name
                    if not donor and has_donor_name and row[col_indices['donor_name']]:
                        full_name = str(row[col_indices['donor_name']]).strip()
                        parts = full_name.split(None, 1)
                        if len(parts) >= 2:
                            donor = Donor.query.filter_by(first_name=parts[0], last_name=parts[1]).first()
                        elif len(parts) == 1:
                            donor = Donor.query.filter_by(last_name=parts[0]).first()

                    # Try split name
                    if not donor and has_split_name:
                        first = row[col_indices['first_name']]
                        last = row[col_indices['last_name']]
                        if first and last:
                            donor = Donor.query.filter_by(
                                first_name=str(first).strip(),
                                last_name=str(last).strip()
                            ).first()

                    if not donor:
                        # Create donor if we have enough info
                        if has_split_name and row[col_indices['first_name']] and row[col_indices['last_name']]:
                            donor = Donor(
                                first_name=str(row[col_indices['first_name']]).strip(),
                                last_name=str(row[col_indices['last_name']]).strip(),
                                email=str(row[col_indices['email']]).strip() if has_email and row[col_indices['email']] else None
                            )
                            db.session.add(donor)
                            db.session.flush()  # Get the ID
                        elif has_donor_name and row[col_indices['donor_name']]:
                            full_name = str(row[col_indices['donor_name']]).strip()
                            parts = full_name.split(None, 1)
                            donor = Donor(
                                first_name=parts[0] if parts else 'Unknown',
                                last_name=parts[1] if len(parts) > 1 else 'Donor',
                                email=str(row[col_indices['email']]).strip() if has_email and row[col_indices['email']] else None
                            )
                            db.session.add(donor)
                            db.session.flush()
                        else:
                            errors.append(f'Row {row_num}: Could not identify donor')
                            continue

                    # Parse date
                    date_val = row[col_indices['date']] if col_indices.get('date') is not None else None
                    donation_date = date.today()
                    if date_val:
                        if isinstance(date_val, datetime):
                            donation_date = date_val.date()
                        elif isinstance(date_val, date):
                            donation_date = date_val
                        else:
                            try:
                                donation_date = datetime.strptime(str(date_val).strip(), '%Y-%m-%d').date()
                            except ValueError:
                                try:
                                    donation_date = datetime.strptime(str(date_val).strip(), '%m/%d/%Y').date()
                                except ValueError:
                                    pass  # Use today's date

                    donation = Donation(
                        donor_id=donor.id,
                        amount=amount,
                        date=donation_date,
                        donation_type=str(row[col_indices['type']]).strip().lower() if col_indices.get('type') is not None and row[col_indices['type']] else 'one-time',
                        campaign=str(row[col_indices['campaign']]).strip() if col_indices.get('campaign') is not None and row[col_indices['campaign']] else None,
                        notes=str(row[col_indices['notes']]).strip() if col_indices.get('notes') is not None and row[col_indices['notes']] else None
                    )
                    db.session.add(donation)
                    imported += 1

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')

            db.session.commit()

            msg = f'Successfully imported {imported} donations.'
            if skipped:
                msg += f' Skipped {skipped} empty rows.'
            if errors:
                msg += f' {len(errors)} errors.'
            flash(msg, 'success' if imported > 0 else 'warning')

            if errors and len(errors) <= 10:
                for err in errors:
                    flash(err, 'warning')

            return redirect(url_for('donations'))

        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'danger')

    return render_template('import_donations.html', form=form)


# =============================================================================
# ADMIN - USER MANAGEMENT
# =============================================================================

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.order_by(User.name).all()
    return render_template('admin_users.html', users=users, User=User)


@app.route('/admin/users/<int:user_id>/role', methods=['POST'])
@login_required
@admin_required
def update_user_role(user_id):
    user = User.query.get_or_404(user_id)
    new_role = request.form.get('role')

    if new_role not in User.ROLES:
        flash('Invalid role.', 'danger')
        return redirect(url_for('admin_users'))

    # Prevent removing the last admin
    if user.role == User.ROLE_ADMIN and new_role != User.ROLE_ADMIN:
        admin_count = User.query.filter_by(role=User.ROLE_ADMIN).count()
        if admin_count <= 1:
            flash('Cannot remove the last admin. Promote another user to admin first.', 'danger')
            return redirect(url_for('admin_users'))

    old_role = user.role
    user.role = new_role
    db.session.commit()
    AuditLog.log('update_role', 'user', user.id, f'Changed role from {old_role} to {new_role} for {user.email}')
    flash(f'Updated role for {user.name} to {new_role}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/audit-log')
@login_required
@admin_required
def admin_audit_log():
    page = request.args.get('page', 1, type=int)
    action_filter = request.args.get('action', '')
    user_filter = request.args.get('user_id', '', type=str)

    query = AuditLog.query

    if action_filter:
        query = query.filter(AuditLog.action == action_filter)
    if user_filter:
        query = query.filter(AuditLog.user_id == int(user_filter))

    logs = query.order_by(AuditLog.timestamp.desc()).paginate(
        page=page, per_page=50, error_out=False
    )

    # Get unique actions for filter dropdown
    actions = db.session.query(AuditLog.action).distinct().all()
    actions = [a[0] for a in actions]

    users = User.query.order_by(User.name).all()

    return render_template('admin_audit_log.html',
        logs=logs,
        actions=actions,
        users=users,
        filters={'action': action_filter, 'user_id': user_filter}
    )


# =============================================================================
# INITIALIZE DATABASE
# =============================================================================

def init_db():
    with app.app_context():
        db.create_all()
        print("Database initialized!")


# Initialize database tables on import (for gunicorn/production)
with app.app_context():
    db.create_all()


if __name__ == '__main__':
    app.run(debug=True)
